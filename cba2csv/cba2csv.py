import os
import sys
import argparse

import csv
import openpyxl  # https://openpyxl.readthedocs.io/en/stable/
import pandas

import re

import datetime
import calendar

import logging
import coloredlogs

from config import *

LOGGING_LEVEL = 'INFO'
LOGGING_FMT = '%(asctime)s %(levelname)s %(message)s'



def month_str_to_number(string):
    m = {
        'jan': 1,
        'feb': 2,
        'mar': 3,
        'apr': 4,
        'may': 5,
        'jun': 6,
        'jul': 7,
        'aug': 8,
        'sep': 9,
        'oct': 10,
        'nov': 11,
        'dec': 12
    }
    s = string.strip()[:3].lower()

    try:
        out = m[s]
        return out
    except:
        raise ValueError('Not a month')





def write_games_csv(games, csv_header, filename):
    print(filename)
    with open(filename, 'w') as f:
        writer = csv.DictWriter(
            f, fieldnames=csv_header, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(games)


def extract_cba_games(coburg_sheet, game_date=None, year=None, game_length_min=40):
    """Builds a list of dictionaries, each dictionary representing a game from a  Coburg Excel sheet

    Args:
        coburg_sheet (str): filename of the Excel sheet sent by Coburg
        game_date (date): the day of the game, if any (otherwise extract from sheet)
        game_length_min (int, optional): how long is each game. Defaults to 40.

    Returns:
        [list(dict)]: a list of dictionaries, each dictionary representing a game with all its info.
                    The dictionary is closed to what TeamAPP needs but not quite there
            e.g.,
            {'start_date': datetime.date(2021, 8, 7), 'end_date': datetime.date(2021, 8, 7), 'start_time': datetime.time(8, 30), 'end_time': datetime.time(9, 10), 'location': 'Coburg Court 2', 'team1': 'Magic Gold', 'team2': 'Piranhas Neon', 'league': 'U9 Mixed'}
    """
    def get_team_id(league):
        """Build the league name

        E.g., if league = "u/13 girls Div3"  it will return "U13 Girls"

        Args:
            league (str): name of teh league as appears in spreadsheet

        Returns:
            str: standarized league string
        """
        # league: u/15 boys Div 1 or  u/13 girls Div3 or u/9 Mixed
        age = re.search("\d+", league.split()[0]).group(0) # extract the age number from u/13 or u16

        gender = "UNKNOWN" # default gender as of 2022
        if "G" in league.upper():
            gender = "Girls"
        if "B" in league.upper():
            gender = "Boys"
        if "M" in league.upper():
            gender = "Mixed"

        if gender == "UNKNOWN":
            gender = "Boys"
            logging.warning(
                f"Could not extract gender on text: {league} - Assuming {gender.upper()}")

        return f"U{age} {gender}"

    if game_date is None:
        # First line of sheet contains the day of the game as two strings ("7th August" and "Winter 2021"
        # Store into date which will then be added the actual time
        day_str = coburg_sheet["C1"].value.strip()   # cell C1: "7th August"
        day_no_str = re.search("\d+", day_str).group(0)
        day_no = int(day_no_str)
        month_name = day_str.split(" ")[1]
        month_no = month_str_to_number((month_name))
        # game_date = datetime.datetime.strptime(f"{year}-{month_no}-{day_no}", "%Y-%m-%d")
        game_date = datetime.date(year, month_no, day_no)

    if year is None:
        # Parse first row to get year if possible
        row1 = list(coburg_sheet.iter_rows(min_row=1, max_row=1, max_col=10))[0]
        for cell in row1:
            if cell.value is not None:
                match = re.search("\d\d\d\d", cell.value.strip())
                if match:
                    year = int(match.group(0))
                    break
        if year is None:
            logging.error(f"Problem, could not extract year in first row of sheet {coburg_sheet}")
            exit(1)

        # Alternative by enumerating cells in row 1 - ugly
        # header_cells = ["E1", "F1", "G1", "H1", "I1", "J1"]
        # for c in header_cells:
        #     if coburg_sheet[c].value is not None:
        #         match = re.search("\d\d\d\d", coburg_sheet[c].value.strip())
        #         if match:
        #             year = int(match.group(0))
        #             break
        logging.warning(f"Extracted year (this is very fragile!): {year}")

    # line 2 has the list of courts available, starting from second colum (first is "TIME")
    #  ['Coburg Court 1', 'Coburg Court 2', 'Coburg Court 3', 'Coburg Court 4'] for Coburg sheet
    # ['Northcote High', 'Oak Park', 'PVG', 'St.Johns'] for external sheet
    courts = [cell.value for cell in coburg_sheet[2]
              [1:] if cell.value != "TIME"]

    # OK so next we collect all games in a list, each game will be a dictionary
    # games start in row three, they are listed by time-slot, each using 3 consecutive lines
    #       columns from B represent different courts.
    #          So lines 3-5 will have the 8:30 slot games, and each colum B-E is a court game
    games_db = []
    games_range = list(coburg_sheet[3:coburg_sheet.max_row])
    no_time_slots = len(games_range) // 3    # no of time-slot in the sheet

    logging.debug(
        f"Processing sheet {coburg_sheet.title} with {no_time_slots} time slot games on courts {courts}")
    for n in range(len(games_range)):
        # the value in the sheet is read as a float because it is 4.5
        time = games_range[n][0].value
        if time is None:
            continue

        # each time slot is built from three consecutives rows
        # contains the teams playing, e.g., "Flames2 vs Jets Blue"
        # each row_x is a tuple of 'openpyxl.cell.cell.Cell'
        row_1 = games_range[n-1]
        # contains the time in the first column, then all blank!
        row_2 = games_range[n]
        # contains the league division, e.g., "u/19 boys Div1"
        row_3 = games_range[n+1]
        # extract hour of games
        hr = int(time)
        if hr < 7:  # shift hr to 234hrs format
            hr += 12
        min = round((time - int(time))*100)  # handles cases like 4.50 and 4.05

        logging.debug(f"=======> Processing games at time {time}")
        for k in range(len(row_1)-1):
            game_dict = dict()
            cell_teams = row_1[k+1]  # class <class 'openpyxl.cell.cell.Cell'>
            cell_league = row_3[k+1] # class <class 'openpyxl.cell.cell.Cell'>
            if cell_teams.value is not None:
                logging.debug(f"Processing game {cell_teams.value} in cell {cell_teams}.")
                # date = datetime.datetime.strptime(f"{year}-{month_no}-{day_no}", "%Y-%m-%d")
                game_dict["cell"] = cell_teams
                game_dict["start_date"] = game_date
                game_dict["end_date"] = game_date
                game_dict["start_time"] = datetime.time(hour=hr, minute=min)
                # calculate end time by adding minutes - datetime.time does not support adding delta directly
                # https://stackoverflow.com/questions/12448592/how-to-add-delta-to-python-datetime-time
                game_dict["end_time"] = (datetime.datetime.combine(datetime.date(
                    1, 1, 1), game_dict["start_time"]) + datetime.timedelta(minutes=game_length_min)).time()

                game_dict["location"] = coburg_sheet[2][k+1].value.strip()
                game_dict["court"] = ""
                if "Coburg Court" in game_dict["location"]:
                    game_dict["court"] = "Court " + re.search("\d", game_dict["location"]).group(0) # extract the court number ("Coburg Court 3")
                    game_dict["location"] = "Coburg Basketball Stadium"
                game_dict["team1"] = cell_teams.value.upper().split("VS")[0].strip().title()     # handles any vs, VS, Vs, vS combinations
                game_dict["team2"] = cell_teams.value.upper().split("VS")[1].strip().title()
                game_dict["league"] = get_team_id(cell_league.value)
                logging.debug(f"\tGame {cell_teams.value} in cell {cell_teams} processed successfully.")
                games_db.append(game_dict)

    return games_db

# Extract the data for venue_key from database of venues in VENUE_INFO constant
def get_venue_info(venue_key):
    if isinstance(VENUES_INFO[venue_key], str):
        return VENUES_INFO[VENUES_INFO[venue_key]]
    else:   # in case it is a re-naming pointing to the real one in the dictionary
        return VENUES_INFO[venue_key]


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="""
        Convert CBA game XLSX timesheet to TeamAPP event/schedule import CSV file.\n

        Neeeds a 2007-Office 265 .xlsx format spreadsheet

        E.g.,:

        $ python cob2csv.py Saturday\ 6th\ November\ 2021\ -Grading\ 1\ updated.xlsx --id "Grading 1"
        """
    )
    parser.add_argument(
        dest='COBURG_SHEET',
        type=str,
        help='The Excel sheet Coburg sends with games.'
    )
    parser.add_argument(
        dest='CLUB',
        type=str,
        help='Club to extract games.',
    )
    parser.add_argument(
        '--id',
        help='Identification of the schedule (e.g., "Grading 1" or "Game 4").'
    )
    parser.add_argument(
        '--description',
        help='Filename with alternative description text to use.'
    )
    parser.add_argument(
        '--no-rsvp',
        action="store_true",
        help='Do not set the RSVP option.',
    )
    parser.add_argument(
        '--no-comments',
        action="store_true",
        help='Do not set the comments option.',
    )
    parser.add_argument(
        '--no-attendance',
        action="store_true",
        help='Do not set the attendance option.',
    )
    parser.add_argument(
        '--no-duty',
        action="store_true",
        help='Do not set the duty roster option.',
    )
    parser.add_argument(
        '--ticketing',
        action="store_true",
        help='Do not set the ticketing option.',
    )
    parser.add_argument(
        '--debug',
        action="store_true",
        default=False,
        help='Show debugging info while processing games (default: %(default)s).',
    )

    # we could also use vars(parser.parse_args()) to make args a dictionary args['<option>']
    args = parser.parse_args()
    print(args)

    option_rsvp = 1 if not args.no_rsvp else 0
    option_comments = 1 if not args.no_comments else 0
    option_attendance = 1 if not args.no_attendance else 0
    option_duty = 1 if not args.no_duty else 0
    option_ticketing = 1 if args.ticketing else 0

    # Set format and level of debug
    coloredlogs.install(level="DEBUG" if args.debug else LOGGING_LEVEL, fmt=LOGGING_FMT)

    if not os.path.exists(args.COBURG_SHEET):
        logging.error(f"File {args.COBURG_SHEET} does not exists!")
        exit(1)

    if args.description is not None:
        try:
            with open(args.description, "r") as f:
                DESCRIPTION = f.read()
        except Exception as e:
            logging.error(f"Unable to read description content from file {args.description}: {e}")

    today = datetime.date.today() #reference point. 
    saturday = today + datetime.timedelta((calendar.SATURDAY-today.weekday()) % 7 )
    logging.info(f"The upcoming Saturday is: {saturday}")


    games_club = []  # these is where all teh CBA games will be collected, each as a dict
    try:    # try to read as a PLAYHQ csv if possible
        with open(args.COBURG_SHEET, newline='') as f: 
            playhq_dict = csv.DictReader(f)

            # filter just the upcoming saturday
            for g in playhq_dict:
                date_game = datetime.datetime.strptime(g['game date'], "%d/%m/%Y").date()
                if date_game == saturday and (args.CLUB in g['team a'] or args.CLUB in g['team b']):
                    game = {}
                    game['team1'] = g['team a'].removeprefix(f"{args.CLUB} ")   # "Magic U16 Girls Gold" -- "U16 Girls Gold"
                    game['team2'] = g['team b'].removeprefix(f"{args.CLUB} ")
                    game['team1'] = g['team a']
                    game['team2'] = g['team b']

                    game['start_date'] = datetime.datetime.strptime(g['game date'], "%d/%m/%Y").date()
                    game['end_date'] = game['start_date']
                    game['start_time'] = datetime.datetime.strptime(g['time'], "%H:%M:%S").time()
                    game["end_time"] = (datetime.datetime.combine(datetime.date(
                    1, 1, 1), game['start_time']) + datetime.timedelta(minutes=GAME_LENGTH_MIN)).time()

                    game['location'] = g['venue']
                    game['court'] = g['playing surface']
                    game['league'] = f"{g['grade'].split()[1]} {g['grade'].split()[2]}" # 'Saturday U16 Girls Division 1/2' --> 'U16 Girls'

                    games_club.append(game)

            logging.info("Finished extracting club games from PLAYHQ CSV")
    except:     # ok, failed to read playhq csv, maybe it is a CBA spreadsheet, let's try that...
        logging.warning(f"File {args.COBURG_SHEET} not a PlayHQ CSV file, trying CBA spreadsheet next...")
        #  0.Load the whole XLSX file
        wb = openpyxl.load_workbook(args.COBURG_SHEET)
        logging.info(
            f"Worksheet {args.COBURG_SHEET} has the following sheets: {wb.sheetnames}")

        coburg_sheet = wb[wb.sheetnames[0]]
        # external_sheet = wb[wb.sheetnames[1]]

        # Collect ALL the games in the worksheet, ALL of them regardless of teams
        games_cba = []
        for sheet in wb:
            logging.info(f"Processing sheet: {sheet.title}")
            games_cba += extract_cba_games(sheet, game_date=saturday, game_length_min=GAME_LENGTH_MIN)
        logging.info(f"Extracted {len(games_cba)} total games.")

        # Project on the team of the club only (args.CLUB) - in our case Magics
        # games_club = [game for game in games_cba if args.CLUB.upper(
        # ) in game['team1'].upper() or args.CLUB.upper() in game['team2'].upper()]

        games_club= []
        for game in games_cba:
            if args.CLUB.upper() in game['team1'].upper() or args.CLUB.upper() in game['team2'].upper():
                games_club.append(game)
                logging.debug(f"This is a club ({args.CLUB}) game: {game['team1']} vs {game['team2']} from {game['cell']}")
            else:
                logging.debug(f"This is NOT a club ({args.CLUB}) game: {game['team1']} vs {game['team2']} from {game['cell']}")




    # OK at this point we have all relevant club games in games_club (either from playhq csv or CBA xls)
    # Next we do a several processings of the games to build the CSV for TeamAPP

    # 1. Report number of club games
    logging.info(f"Filtered {len(games_club)} for team: {args.CLUB}.")
    if len(games_club) == 0:
        logging.warning(f"No games extracted for club \"{args.CLUB}\". Correct spelling? Stopping...")
        exit(1)


    # 2. Extract date of first game
    date_games = games_club[0]['start_date'] 

    # 3. Extract Club team and opponent team; build new games if both are Club's teams
    games_swap = []  # collect new games where both teams1 an teams2 are --club games
    for game in games_club:
        # game['league'] = "U12 Boys"
        # game['team1'] = "Magic Gold"
        if args.CLUB.upper() in game['team1'].upper():
            # build "U12 Boys Gold" from "Magic U12 Boys Gold"
            game["team_name"] = f"{game['league']} {game['team1'].split()[-1]}".title()
            game["opponent"] = game['team2']

            # both team1 and team2 are args.CLUB! Make a duplicate
            if args.CLUB.upper() in game['team2'].upper():
                game2 = game.copy()  # make a fresh copy of the game
                game2["team_name"] = f"{game2['league']} {game2['team2'].split()[-1]}".title()
                game2["opponent"] = game2['team1']
                # new entry for team2 as both teams are magic
                games_swap.append(game2)
        elif args.CLUB.upper() in game['team2'].upper():    # club only as team2
            game["team_name"] = f"{game['league']} {game['team2'].split()[-1]}".title()
            game["opponent"] = game['team1']

        game["opponent"] = re.sub(r" U\d* ", " ", game['opponent']) # remove the UXX in opponent
    games_club += games_swap

    # 4. Add teamapp options to each game record
    for game in games_club:
        if not game['team_name'] in CLUB_TEAMS:
            continue

        logging.info(f"Found game for club team {game['team_name']}: against {game['opponent']} at {game['start_time']} in {game['location']}")

        # U12 Boys Gold - Grading 1
        game["event_name"] = game["team_name"] + \
            (f" - {args.id}" if args.id is not None else "")
        game["access_groups"] = game["team_name"]

        for venue_key in VENUES_INFO.keys():
            if venue_key in game['location']:
                venue_data = get_venue_info(venue_key)

                game['location'] = venue_data[1]  # address of venue
                # e.g., Coburg Stadium
                game['venue'] = venue_data[0]
                game['addr_tips'] = ""
                if venue_data[2] is not None:
                    game['addr_tips'] = f" ({venue_data[2]})"

        # could't fill the venue key in game: venue data not found!
        if "venue" not in game:
            logging.error(f"Error, no venue matched for {game}")
            sys.exit(1)

        # Build the whole description of the game
        game["description"] = DESCRIPTION.format(
            opponent=game["opponent"],
            venue=game['venue'],
            court=game['court'],
            address=game['location'],
            address_tips=game['addr_tips'])
        game["rsvp"] = option_rsvp
        game["comments"] = option_comments
        game["attendance_tracking"] = option_ticketing
        game["duty_roster"] = option_duty
        game["ticketing"] = option_ticketing
        game["reference_id"] = ""

    # 5. Add BYE games
    teams_playing = [x['team_name'] for x in games_club]
    teams_bye = CLUB_TEAMS.difference(teams_playing)

    if not teams_bye:
        logging.info("No teams have a BYE game; all are playing!")
    else:
        logging.info("Teams having BYE (i.e., have not found any game for club team):")

        for team in teams_bye:
            print(f"\t {team}")

            game = {}
            game["team_name"] = team
            game["access_groups"] = game["team_name"]
            game["start_date"] = date_games
            game["end_date"] = date_games
            game["start_time"] = datetime.time(hour=11, minute=00)
            game["end_time"] =datetime.time(hour=12, minute=00)

            game["event_name"] = f"{game['team_name']} - BYE"

            game["opponent"] = ""
            game['location'] = ""
            game['venue'] = ""
            game["description"] = DESCRIPTION_BYE

            game["rsvp"] = 0
            game["comments"] = 0
            game["attendance_tracking"] = 0
            game["duty_roster"] = 0
            game["ticketing"] = 0
            game["reference_id"] = ""

            games_club.append(game)

    #  6. Check for games on teams that do no exist in the club; something is wrong!
    for game in games_club:
        if not game['team_name'] in CLUB_TEAMS:
            logging.error(f"Found game for NON-EXISTENT CLUB team {game['team_name']}: against {game['opponent']} at {game['start_time']} in {game['location']}")
            logging.error(f"Check if this would correspond to some BYE team above. Fix the sheet and re-run script")
            sys.exit(1)


    #  7. Report all games if in DEBUG level
    ## This is even too much, uncomment if you really want it
    # if coloredlogs.get_level() == coloredlogs.level_to_number("DEBUG"):
    #     print("========== TEAMS EXTRACTED FOR TEAMAPP ===============")
    #     for game in games_club:
    #         print(game)
    #     print("======================================================")


    #  8. Report Games
    print("=========================================")
    print("GAMES EXTRACTED")
    print("=========================================")
    for g in sorted(games_club, key=lambda x: x['team_name']):
        # print(g)
        if g['venue'] != "": # not a bye
            print(f"{g['team_name']} vs {g['opponent']} @ {g['venue']} - {g['start_time']}"  )
        else:
            print(f"{g['team_name']} has a BYE")
    print("=========================================")

    #  9. Produce CSV for TeamAPP
    schedule_file_path = os.path.split(args.COBURG_SHEET)[0]
    schedule_file_name = os.path.splitext(os.path.split(args.COBURG_SHEET)[1])[0]   # without extension

    csv_header_schedule = ["event_name",  "team_name", "start_date", "end_date", "start_time", "end_time", "description",
                           "location", "access_groups", "rsvp", "comments", "attendance_tracking", "duty_roster", "ticketing", "reference_id"]

    # prefix e.g., "2022-05-14.Magic_Grading-2"
    prefix = f"{date_games}.{args.CLUB}" + (f"_{args.id.replace(' ', '-')}" if args.id is not None else "")

    # write the CSV to import as a SCHEDULE
    schedule_csv = os.path.join(schedule_file_path, f"{prefix}_SCHEDULE-{schedule_file_name}.csv")
    write_games_csv(games_club, csv_header_schedule, schedule_csv)
    logging.info(
        f"CSV schedule file for TeamAPP produced for team {args.CLUB} ({len(games_club)} games: {schedule_csv}.")

    # write the CSV to import as an EVENT
    csv_header_events = csv_header_schedule
    csv_header_events.remove("team_name")
    events_csv = os.path.join(schedule_file_path, f"{prefix}_EVENT-{schedule_file_name}.csv")
    write_games_csv(games_club, csv_header_events, events_csv)
    logging.info(
        f"CSV event file for TeamAPP produced for team {args.CLUB} ({len(games_club)} games: {events_csv}.")


